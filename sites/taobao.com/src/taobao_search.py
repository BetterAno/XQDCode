"""
淘宝搜索商品采集工具
=====================
用法:
  python src/taobao_search.py                                → 输入关键词，输出浏览器提取脚本
  python src/taobao_search.py --keyword 耳机                  → 指定关键词，输出提取脚本
  python src/taobao_search.py --screenshot data.json          → 自动截图每个商品详情页
  python src/taobao_search.py --to-xlsx data.json             → JSON + 截图 → XLSX（含截图列）

浏览器:
  .cloakbrowser 自定义 Chromium + E:/Projects/ChromeDebug 配置

依赖: playwright openpyxl Pillow
"""

import sys
import io
import os
import json
import datetime
import traceback
import subprocess
import random
import time

# ╔══════════════════════════════════════════════════════════════════╗
# ║                       0. 配置                                    ║
# ╚══════════════════════════════════════════════════════════════════╝

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JS_PATH = os.path.join(BASE_DIR, "extract_search.js")
SCREENSHOTS_DIR = os.path.join(BASE_DIR, "screenshots")

# 浏览器配置（参考 start_chrome.bat）
CHROME_EXE = os.path.expandvars(
    r"%USERPROFILE%\.cloakbrowser\chromium-146.0.7680.177.5\chrome.exe"
)
CHROME_PROFILE = r"E:\Projects\ChromeDebug"

# 截图延时（秒）—— 适当拉长防检测
PAGE_WAIT = (4, 6)       # 详情页加载后等待 (min, max)
BETWEEN_WAIT = (2, 3)    # 商品间间隔

# XLSX 列定义
HEADERS = [
    "序号", "店铺名", "标题", "售价", "地址",
    "销量", "详情页链接", "商品图片", "itemId", "商品详情页截图",
]
COL_WIDTHS = [6, 28, 60, 12, 12, 14, 50, 50, 16, 30]
IMG_WIDTH_PT = 220   # 截图嵌入宽度（点）
IMG_HEIGHT_PT = 155  # 截图嵌入高度（点）


# ╔══════════════════════════════════════════════════════════════════╗
# ║                       1. 工具函数                                ║
# ╚══════════════════════════════════════════════════════════════════╝

def _log(level: str, msg: str):
    """统一日志：带级别前缀，方便排查"""
    prefixes = {"INFO": "  ", "OK": "✅", "WARN": "⚠️", "ERR": "❌", "STEP": "📌"}
    prefix = prefixes.get(level, "  ")
    print(f"{prefix} {msg}", flush=True)


def log_step(msg: str):
    _log("STEP", msg)


def log_ok(msg: str):
    _log("OK", msg)


def log_warn(msg: str):
    _log("WARN", msg)


def log_err(msg: str):
    _log("ERR", msg)


def find_chrome_path() -> str | None:
    """查找 .cloakbrowser 自定义 Chromium。找不到返回 None。"""
    if os.path.exists(CHROME_EXE):
        return CHROME_EXE
    return None


def check_captcha(page) -> bool:
    """
    多维度判断页面是否异常（验证码/登录/风控）。
    返回 True → 已暂停等待用户处理完毕。
    """
    url_lower = page.url.lower()
    reason = ""

    try:
        # ---- 第 0 层：主动等待验证码 DOM 出现（5s 内）----
        # 滑块可能延迟渲染，用 wait_for_selector 等待
        try:
            page.wait_for_selector(
                "#nocaptcha, .captcha-tips, #nc_1_n1z, .sm-pop-inner, .slidetounlock",
                timeout=5000, state="attached",
            )
        except Exception:
            pass  # 没等到，继续后续检查

        # ---- 第 1 层：URL 检查 ----
        is_product_url = ("item.taobao.com" in url_lower
                          or "detail.tmall.com" in url_lower)
        if not is_product_url:
            reason = "URL 不是商品详情页"
            raise Exception(reason)

        if any(kw in url_lower for kw in ["login.", "verify.", "captcha", "punish"]):
            reason = f"URL 包含验证关键词"
            raise Exception(reason)

        # ---- 第 2 层：页面内容量检查（验证码页内容极少）----
        body_len = page.evaluate("() => (document.body?.innerText || '').length")
        if body_len < 150:
            reason = f"页面文本仅 {body_len} 字符（过短，疑似验证码页）"
            raise Exception(reason)

        # ---- 第 3 层：商品页特征检查 ----
        has_product = page.evaluate(r"""
        () => {
          const t = (document.body?.innerText || '').slice(0, 2000).toLowerCase();
          const signals = ['¥','价格','发货','月销量','件','加入购物车','立即购买','店铺','收藏','宝贝','评价','包邮'];
          let n = 0; for (const s of signals) { if (t.includes(s)) n++; }
          return n >= 3;
        }
        """)
        if not has_product:
            reason = "页面缺少商品特征（无价格/销量/店铺）"
            raise Exception(reason)

        # ---- 第 4 层：JS 全量扫描（主页面 + iframe）----
        captcha_found = page.evaluate(r"""
        () => {
          const SEL = [
            // 淘宝/天猫滑块验证码（基于真实 HTML）
            '.captcha-tips','.warnning-text','.sm-pop-inner',
            '#nocaptcha','#nc_1_n1z','#nc_1_wrapper','.nc_wrapper',
            '.nc_scale','.btn_slide','#nc_1__scale_text','.slidetounlock',
            '.nc-lang-cnt','.nc_iconfont','#nc-verify-form',
            // 通用
            '.captcha','.captcha-qrcode','.captcha-uuid-tips',
            '[id*="captcha"]','[class*="captcha"]',
            '[id*="sm_"]','.baxia-dialog','#aliyunCaptcha',
            '[class*="verify"]','#baxia',
          ];
          const search = (doc) => {
            for (const s of SEL) {
              try { const el = doc.querySelector(s); if (el) return true; } catch(e) {}
            }
            const txt = (doc.body?.innerText || '');
            // 滑块文案、验证码提示
            if (/请按住滑块|拖动.*最右边|滑块完成验证|验证以确保正常访问/.test(txt)) return true;
            if (txt.slice(0,600).length < 100) return true; // 极短文本
            return false;
          };
          if (search(document)) return true;
          for (const f of document.querySelectorAll('iframe')) {
            try { if (search(f.contentDocument)) return true; } catch(e) {}
          }
          return false;
        }
        """)
        if captcha_found:
            reason = "DOM 命中验证码元素"
            raise Exception(reason)

        # ---- 第 5 层：标题检查 ----
        title = page.title()
        if any(kw in title for kw in ["验证", "登录", "verify", "login", "安全"]):
            reason = f"标题异常: {title}"
            raise Exception(reason)

        return False

    except Exception as e:
        if not reason:
            reason = str(e).replace("Exception: ", "")[:60]

    print()
    print("  ╔══════════════════════════════════════════════╗")
    print(f"  ║  🛑 页面异常 [{reason}]")
    print("  ║     请在浏览器中完成验证/登录                 ║")
    print("  ║     完成后回到此处按 Enter 继续截图            ║")
    print("  ╚══════════════════════════════════════════════╝")
    input("  >>> 按 Enter 继续...")
    print()
    log_ok("已确认，继续截图...")
    return True


def make_filename(keyword: str) -> str:
    """生成输出文件名: {关键词}_taobao_{日期}.json"""
    date_str = datetime.date.today().isoformat()
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return f"{safe_kw}_taobao_{date_str}.json"


def resolve_json_path(json_path: str) -> str:
    """解析 JSON 文件路径。支持相对路径（补全 .json 后缀）。文件不存在时直接退出。"""
    if not json_path.endswith(".json"):
        json_path += ".json"
    if not os.path.isabs(json_path):
        json_path = os.path.join(BASE_DIR, json_path)
    if not os.path.exists(json_path):
        log_err(f"文件不存在: {json_path}")
        sys.exit(1)
    return json_path


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 1: 生成浏览器提取脚本                      ║
# ╚══════════════════════════════════════════════════════════════════╝

def get_keyword() -> str:
    """从 --keyword 参数或交互输入获取搜索关键词。"""
    if "--keyword" in sys.argv:
        idx = sys.argv.index("--keyword")
        if idx + 1 < len(sys.argv):
            return sys.argv[idx + 1]
    return input("请输入搜索关键词: ").strip()


def step1_generate_js(keyword: str):
    """
    读取 extract_search.js 模板，替换关键词占位符，
    输出完整的 Console 粘贴脚本 + 使用说明。
    """
    with open(JS_PATH, "r", encoding="utf-8") as f:
        js_code = f.read()
    js_code = js_code.replace("__KEYWORD__", keyword)

    filename = make_filename(keyword)

    print()
    log_step(f"关键词: {keyword}")
    log_step(f"输出文件: {filename}")
    print()
    print("  使用步骤:")
    print(f"    1. 浏览器打开 https://s.taobao.com/search?q={keyword}")
    print(f"    2. F12 → Console → 粘贴下方脚本 → 回车")
    print(f"    3. 等待自动滚动加载 → 自动下载 {filename}")
    print(f"    4. 将 {filename} 移到项目根目录")
    print(f"    5. python src/taobao_search.py --screenshot {filename}")
    print(f"    6. python src/taobao_search.py --to-xlsx {filename}")
    print()
    print("-" * 56)
    print("👇 复制以下脚本到浏览器 Console 👇")
    print("-" * 56)
    print(js_code)
    print("-" * 56)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                  Step 2: 商品详情页截图                           ║
# ╚══════════════════════════════════════════════════════════════════╝

def step2_screenshot(json_path: str):
    """
    读取 JSON 商品列表，启动浏览器逐条打开详情页截图。

    浏览器: .cloakbrowser Chromium + E:/Projects/ChromeDebug 配置
    自动先关旧 Chrome 防 Profile 占用。

    执行流程:
      1. 关闭所有 Chrome → 启动浏览器（自带登录态）
      2. 逐条打开商品详情页，等待渲染 → 截图
      3. 自动跳过已有截图（支持断点续传）
    """
    # --- 2.1 校验依赖 & 路径 ---
    try:
        from playwright.sync_api import sync_playwright  # noqa: F811
    except ImportError:
        log_err("缺少 playwright，请执行: pip install playwright")
        sys.exit(1)

    exe_path = find_chrome_path()
    if not exe_path or not os.path.exists(exe_path):
        log_err(f"Chromium 不存在: {CHROME_EXE}")
        log_err("请确认 .cloakbrowser 已安装")
        sys.exit(1)

    log_ok(f"浏览器: {exe_path}")
    log_ok(f"用户目录: {CHROME_PROFILE}")

    # --- 2.2 加载数据 ---
    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        items = json.load(f)
    log_step(f"共 {len(items)} 个商品待截图")

    # --- 2.3 增量跳过 ---
    os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
    done_ids = {f.replace(".png", "") for f in os.listdir(SCREENSHOTS_DIR) if f.endswith(".png")}
    pending = [it for it in items if it["itemId"] not in done_ids]
    skipped = len(items) - len(pending)

    if skipped:
        log_ok(f"跳过 {skipped} 个已有截图，待处理 {len(pending)} 个")
    if not pending:
        log_ok("所有商品已有截图，无需处理")
        return

    # --- 2.4 启动浏览器（先关旧的防 Profile 占用） ---
    log_step("关闭旧的 Chrome 进程...")
    for _ in range(3):
        subprocess.run(["taskkill", "/f", "/im", "chrome.exe"],
                       capture_output=True, timeout=10)
        subprocess.run(["taskkill", "/f", "/im", "chromium.exe"],
                       capture_output=True, timeout=10)

    success = 0
    fail = 0

    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=CHROME_PROFILE,
                executable_path=exe_path,
                headless=False,
                viewport={"width": 1280, "height": 900},
                args=["--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            log_err("浏览器启动失败！")
            log_err(f"  1. 请确认已关闭所有 Chrome 窗口")
            log_err(f"  2. 浏览器路径: {exe_path}")
            log_err(f"  3. Profile 目录: {CHROME_PROFILE}")
            log_err(f"\n原始错误: {traceback.format_exc().splitlines()[-1]}")
            sys.exit(1)

        page = context.pages[0] if context.pages else context.new_page()
        log_ok("浏览器已启动，开始截图...")
        print()

        # --- 2.5 逐个截图 ---
        for idx, item in enumerate(pending, 1):
            item_id = item["itemId"]
            detail_url = item.get("detailUrl", "")

            if not detail_url:
                log_warn(f"[{idx}/{len(pending)}] 无详情链接，跳过: {item_id}")
                fail += 1
                continue

            snapshot_path = os.path.join(SCREENSHOTS_DIR, f"{item_id}.png")

            try:
                print(f"  [{idx}/{len(pending)}] {item_id} ... ",
                      end="", flush=True)

                page.goto(detail_url, timeout=30000, wait_until="domcontentloaded")

                # 验证码检测（内置 5s 轮询等待 DOM）
                if check_captcha(page):
                    page.wait_for_timeout(2000)  # 用户处理后等页面恢复

                # 等待页面渲染
                wait_s = random.randint(*PAGE_WAIT)
                page.wait_for_timeout(wait_s * 1000)

                page.screenshot(path=snapshot_path, full_page=False)
                print(f"✓ ({wait_s}s)", flush=True)
                success += 1

                # 商品间停顿
                if idx < len(pending):
                    time.sleep(random.uniform(*BETWEEN_WAIT))
            except Exception:
                print("✗", flush=True)
                log_warn(f"[{idx}/{len(pending)}] {item_id} 失败: "
                         f"{traceback.format_exc().splitlines()[-1]}")
                fail += 1

        context.close()

    # --- 2.6 汇总 ---
    total = success + skipped + fail
    print()
    log_ok(f"截图完成: 新增 {success}, 跳过 {skipped}, 失败 {fail}, 共 {total}/{len(items)}")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 3: JSON → XLSX                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def step3_to_xlsx(json_path: str):
    """
    将搜索结果 JSON + 截图合并为一个 XLSX 文件。
    截图以图片形式嵌入到「商品详情页截图」列。
    """
    # --- 3.1 校验依赖 ---
    try:
        from openpyxl import Workbook  # noqa: F811
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # noqa: F811
        from openpyxl.drawing.image import Image as XlImage  # noqa: F811
        from openpyxl.utils import get_column_letter  # noqa: F811
    except ImportError:
        log_err("缺少 openpyxl，请执行: pip install openpyxl Pillow")
        sys.exit(1)

    # --- 3.2 加载数据 ---
    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        items = json.load(f)

    if not items:
        log_err("JSON 文件为空或无数据")
        return

    log_step(f"数据 {len(items)} 条 → 生成 XLSX ...")

    # --- 3.3 写入工作簿 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "淘宝搜索结果"

    # 表头样式
    hdr_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 列宽
    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # 数据行
    data_font = Font(name="微软雅黑", size=10)
    screenshot_col = len(HEADERS)  # 最后一列 = 截图列号
    img_ok = 0
    img_miss = 0

    for row_idx, item in enumerate(items, 2):
        row_data = [
            row_idx - 1,
            item.get("shopName", ""),
            item.get("title", ""),
            item.get("price", ""),
            item.get("location", ""),
            item.get("sales", ""),
            item.get("detailUrl", ""),
            item.get("image", ""),
            item.get("itemId", ""),
            "",  # 截图列——后面嵌入图片
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            align_h = "center" if col in (1, 4, 5, 6, 9) else "left"
            cell.alignment = Alignment(horizontal=align_h, vertical="center",
                                       wrap_text=(col in (2, 3, 7, 8)))

        # 行高（预留给截图）
        ws.row_dimensions[row_idx].height = IMG_HEIGHT_PT + 8

        # --- 嵌入截图 ---
        item_id = item.get("itemId", "")
        img_path = os.path.join(SCREENSHOTS_DIR, f"{item_id}.png")

        if os.path.exists(img_path):
            try:
                img = XlImage(img_path)
                img.width = IMG_WIDTH_PT
                img.height = IMG_HEIGHT_PT
                ws.add_image(img, f"{get_column_letter(screenshot_col)}{row_idx}")
                img_ok += 1
            except Exception:
                log_warn(f"嵌入截图失败: {item_id}")
                ws.cell(row=row_idx, column=screenshot_col, value="(截图加载失败)")
                img_miss += 1
        else:
            ws.cell(row=row_idx, column=screenshot_col, value="(无截图)")
            img_miss += 1

    # --- 3.4 保存 ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(items) + 1}"

    out_path = json_path.rsplit(".", 1)[0] + ".xlsx"
    wb.save(out_path)

    print()
    log_ok(f"XLSX 已保存: {out_path}")
    log_ok(f"数据 {len(items)} 条 | 截图嵌入 {img_ok}/{len(items)} 张")
    if img_miss:
        log_warn(f"{img_miss} 条缺少截图（需先运行 --screenshot）")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                         主入口                                   ║
# ╚══════════════════════════════════════════════════════════════════╝

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    print("=" * 56)
    print("       淘宝搜索商品采集工具")
    print("=" * 56)

    args = sys.argv[1:]

    if len(args) >= 2 and args[0] == "--screenshot":
        step2_screenshot(args[1])

    elif len(args) >= 2 and args[0] == "--to-xlsx":
        step3_to_xlsx(args[1])

    else:
        keyword = get_keyword()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)
        step1_generate_js(keyword)


if __name__ == "__main__":
    main()
