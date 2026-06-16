"""
闲鱼搜索商品采集工具
=====================
三种运行模式:
  模式1: python src/xianyu_search.py 耳机                           → 指定关键词，输出浏览器提取脚本
  模式2: python src/xianyu_search.py --from-json data.json            → 根据JSON文件：截图+转XLSX
  模式3: python src/xianyu_search.py --auto 耳机                      → 一键自动采集（打开浏览器→注入JS→截图→XLSX）

可选参数:
  --pages N         指定采集页数（默认5）
  --no-captcha      关闭验证码暂停检测（默认开启：遇到验证码会暂停等待处理）

浏览器:
  标准 Chrome + ChromeDebug 独立缓存目录（不影响用户正在使用的 Chrome）

依赖: playwright openpyxl Pillow
"""

import sys
import io
import os
import json
import datetime
import traceback
import subprocess
import random
import time

# ╔══════════════════════════════════════════════════════════════════╗
# ║                       0. 配置                                    ║
# ╚══════════════════════════════════════════════════════════════════╝

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
JS_PATH = os.path.join(BASE_DIR, "extract_search.js")
SCREENSHOTS_DIR = os.path.join(BASE_DIR, "screenshots")

# 浏览器配置（标准 Chrome + 独立 Profile 目录，不影响用户正在使用的 Chrome）
CHROME_EXE = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
CHROME_PROFILE = r"C:\Users\Administrator\Desktop\XQDCode\ChromeDebug"

# 截图延时（秒）—— 闲鱼反爬较严，适当拉长
PAGE_WAIT = (3, 5)       # 详情页加载后等待 (min, max)
BETWEEN_WAIT = (2, 4)    # 商品间间隔

# XLSX 列定义
HEADERS = [
    "序号", "卖家", "标题", "售价", "想要",
    "详情页链接", "商品图片", "itemId", "商品详情页截图",
]
COL_WIDTHS = [6, 22, 60, 12, 12, 50, 50, 18, 30]
IMG_WIDTH_PT = 220   # 截图嵌入宽度（点）
IMG_HEIGHT_PT = 155  # 截图嵌入高度（点）

DEFAULT_PAGES = 5     # 默认采集页数


import urllib.request
import io as _io

# ╔══════════════════════════════════════════════════════════════════╗
# ║                       1. 工具函数                                ║
# ╚══════════════════════════════════════════════════════════════════╝

def _log(level: str, msg: str):
    """统一日志：带级别前缀，方便排查"""
    prefixes = {"INFO": "  ", "OK": "✅", "WARN": "⚠️", "ERR": "❌", "STEP": "📌"}
    prefix = prefixes.get(level, "  ")
    print(f"{prefix} {msg}", flush=True)


def log_step(msg: str):
    _log("STEP", msg)


def log_ok(msg: str):
    _log("OK", msg)


def log_warn(msg: str):
    _log("WARN", msg)


def log_err(msg: str):
    _log("ERR", msg)


def find_chrome_path() -> str | None:
    """查找 Chrome。找不到返回 None。"""
    if os.path.exists(CHROME_EXE):
        return CHROME_EXE
    alt = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    if os.path.exists(alt):
        return alt
    return None


def check_captcha(page) -> bool:
    """
    多维度判断闲鱼页面是否异常（验证码/登录/风控）。
    返回 True → 已暂停等待用户处理完毕。
    """
    url_lower = page.url.lower()
    reason = ""

    try:
        # ---- 第 0 层：主动等待验证码 DOM 出现（5s 内）----
        try:
            page.wait_for_selector(
                ".captcha_verify_container, .verify-captcha, "
                ".login-modal, .goofish-login, "
                "#nocaptcha, #nc_1_n1z, .captcha-box, .sm-pop-inner, "
                ".slidetounlock, .bx_captcha_iframe, #baxia-dialog",
                timeout=5000, state="attached",
            )
        except Exception:
            pass

        # ---- 第 1 层：URL 检查 ----
        is_item_url = "goofish.com/item" in url_lower
        if not is_item_url:
            reason = "URL 不是商品详情页"
            raise Exception(reason)

        if any(kw in url_lower for kw in ["login.", "verify.", "captcha", "punish", "sec-verify"]):
            reason = "URL 包含验证关键词"
            raise Exception(reason)

        # ---- 第 2 层：页面内容量检查 ----
        body_len = page.evaluate("() => (document.body?.innerText || '').length")
        if body_len < 150:
            reason = f"页面文本仅 {body_len} 字符（过短，疑似验证码页）"
            raise Exception(reason)

        # ---- 第 3 层：商品页特征检查 ----
        has_product = page.evaluate(r"""
        () => {
          const t = (document.body?.innerText || '').slice(0, 2000).toLowerCase();
          const signals = ['¥','价格','包邮','想要','卖家','宝贝','闲鱼','担保交易','聊一聊'];
          let n = 0; for (const s of signals) { if (t.includes(s)) n++; }
          return n >= 3;
        }
        """)
        if not has_product:
            reason = "页面缺少商品特征"
            raise Exception(reason)

        # ---- 第 4 层：JS 全量扫描（主页面 + iframe）----
        captcha_found = page.evaluate(r"""
        () => {
          const SEL = [
            '.captcha-tips','.warnning-text','.sm-pop-inner',
            '#nocaptcha','#nc_1_n1z','.nc_wrapper',
            '.nc_scale','.btn_slide','.slidetounlock',
            '.captcha_verify_container','.verify-captcha',
            '.captcha-box','#captcha_container',
            '.login-modal','.goofish-login','.login-mask',
            '.captcha','.captcha-qrcode',
            '[id*="captcha"]','[class*="captcha"]',
            '[class*="verify"]','#baxia-dialog',
          ];
          const search = (doc) => {
            for (const s of SEL) {
              try { const el = doc.querySelector(s); if (el) return true; } catch(e) {}
            }
            const txt = (doc.body?.innerText || '');
            if (/请按住滑块|拖动.*最右边|滑块完成验证|验证以确保正常访问|请先登录|登录后即可/.test(txt)) return true;
            if (txt.slice(0,600).length < 100) return true;
            return false;
          };
          if (search(document)) return true;
          for (const f of document.querySelectorAll('iframe')) {
            try { if (search(f.contentDocument)) return true; } catch(e) {}
          }
          return false;
        }
        """)
        if captcha_found:
            reason = "DOM 命中验证码元素"
            raise Exception(reason)

        # ---- 第 5 层：标题检查 ----
        title = page.title()
        if any(kw in title for kw in ["验证", "登录", "verify", "login", "安全", "滑块"]):
            reason = f"标题异常: {title}"
            raise Exception(reason)

        return False

    except Exception as e:
        if not reason:
            reason = str(e).replace("Exception: ", "")[:60]

    print()
    print("  ╔══════════════════════════════════════════════╗")
    print(f"  ║  🛑 页面异常 [{reason}]")
    print("  ║     请在浏览器中完成验证/登录                 ║")
    print("  ║     完成后回到此处按 Enter 继续截图            ║")
    print("  ╚══════════════════════════════════════════════╝")
    input("  >>> 按 Enter 继续...")
    print()
    log_ok("已确认，继续截图...")
    return True


def make_filename(keyword: str) -> str:
    """生成输出文件名: {关键词}_xianyu_{日期}.json"""
    date_str = datetime.date.today().isoformat()
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return f"{safe_kw}_xianyu_{date_str}.json"


def resolve_json_path(json_path: str) -> str:
    """解析 JSON 文件路径。支持相对路径（补全 .json 后缀）。文件不存在时直接退出。"""
    if not json_path.endswith(".json"):
        json_path += ".json"
    if not os.path.isabs(json_path):
        json_path = os.path.join(BASE_DIR, json_path)
    if not os.path.exists(json_path):
        log_err(f"文件不存在: {json_path}")
        sys.exit(1)
    return json_path


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 1: 生成浏览器提取脚本                      ║
# ╚══════════════════════════════════════════════════════════════════╝

def get_keyword() -> str:
    """从 --keyword 参数或交互输入获取搜索关键词。"""
    if "--keyword" in sys.argv:
        idx = sys.argv.index("--keyword")
        if idx + 1 < len(sys.argv):
            return sys.argv[idx + 1]
    return input("请输入搜索关键词: ").strip()


def step1_generate_js(keyword: str, max_pages: int = DEFAULT_PAGES):
    """
    读取 extract_search.js 模板，替换关键词和页数占位符，
    输出完整的 Console 粘贴脚本 + 使用说明。
    """
    with open(JS_PATH, "r", encoding="utf-8") as f:
        js_code = f.read()
    js_code = js_code.replace("__KEYWORD__", keyword)
    js_code = js_code.replace("__MAX_PAGES__", str(max_pages))

    filename = make_filename(keyword)

    print()
    log_step(f"关键词: {keyword}")
    log_step(f"输出文件: {filename}")
    print()
    print("  使用步骤:")
    print(f"    1. 浏览器打开 https://www.goofish.com/search?q={keyword}")
    print(f"    2. F12 → Console → 粘贴下方脚本 → 回车")
    print(f"    3. 等待自动滚动加载 → 自动下载 {filename}")
    print(f"    4. 将 {filename} 移到项目根目录")
    print(f"    5. python src/xianyu_search.py --from-json {filename}")
    print()
    print("-" * 56)
    print("👇 复制以下脚本到浏览器 Console 👇")
    print("-" * 56)
    print(js_code)
    print("-" * 56)


# ╔══════════════════════════════════════════════════════════════════╗
# ║               Step 0: 一键自动化采集                              ║
# ╚══════════════════════════════════════════════════════════════════╝

def _build_extract_js(keyword: str, max_pages: int = DEFAULT_PAGES) -> str:
    """从 extract_search.js 构建可被 page.evaluate() 执行的提取函数。
    去掉文件下载逻辑，替换为 return items 将数据传回 Python。"""
    with open(JS_PATH, "r", encoding="utf-8") as f:
        js = f.read()
    js = js.replace("__KEYWORD__", keyword)
    js = js.replace("__MAX_PAGES__", str(max_pages))

    # 截断 "  // 下载 JSON" 之后的内容，替换为 return items
    marker = "  // 下载 JSON"
    idx = js.find(marker)
    if idx > 0:
        js = js[:idx] + "  return items;\n})()"

    return js


def step0_auto_collect(keyword: str, max_pages: int = DEFAULT_PAGES,
                       captcha_pause: bool = True) -> str:
    """
    一键自动化：Playwright 打开浏览器 → 搜索 → 注入JS提取数据 → 截图 → XLSX。
    返回 JSON 文件路径。
    """
    try:
        from playwright.sync_api import sync_playwright  # noqa: F811
    except ImportError:
        log_err("缺少 playwright，请执行: pip install playwright")
        sys.exit(1)

    exe_path = find_chrome_path()
    if not exe_path or not os.path.exists(exe_path):
        log_err(f"Chrome 不存在: {CHROME_EXE}")
        sys.exit(1)

    filename = make_filename(keyword)
    json_path = os.path.join(BASE_DIR, filename)
    extract_js = _build_extract_js(keyword, max_pages)

    log_ok(f"浏览器: {exe_path}")
    log_step(f"关键词: {keyword} | 页数: {max_pages} | 验证码暂停: {'开' if captcha_pause else '关'}")
    log_step(f"输出文件: {filename}")

    from urllib.parse import quote
    search_url = f"https://www.goofish.com/search?q={quote(keyword)}"

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=CHROME_PROFILE,
            executable_path=exe_path,
            headless=False,
            viewport={"width": 1280, "height": 900},
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = context.pages[0] if context.pages else context.new_page()
        log_ok("浏览器已启动")

        # 打开搜索页
        log_step(f"打开: {search_url}")
        page.goto(search_url, timeout=30000, wait_until="domcontentloaded")

        # 等待搜索结果出现
        log_step("等待搜索结果加载...")
        try:
            page.wait_for_selector(
                'a[href*="goofish.com/item"] img[class*="feeds-image"]',
                timeout=20000,
            )
        except Exception:
            log_warn("20秒内未检测到搜索结果，可能需验证/登录")
            log_warn("请在浏览器中完成验证后按 Enter 继续...")
            input("  >>> 按 Enter 继续...")

        # 注入提取脚本
        log_step("开始滚动加载并提取数据...")
        items = page.evaluate(extract_js)

        context.close()

    if not items or len(items) == 0:
        log_err("未提取到任何数据！请确认:")
        log_err("  1. 页面已加载搜索结果")
        log_err("  2. 未触发验证码/登录")
        sys.exit(1)

    # 保存 JSON
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    log_ok(f"已保存 {len(items)} 条 → {json_path}")

    # 自动继续截图+XLSX
    print()
    log_step("继续执行截图...")
    step2_screenshot(filename, captcha_pause=captcha_pause)
    print()
    log_step("继续生成 XLSX...")
    step3_to_xlsx(filename)

    return json_path


# ╔══════════════════════════════════════════════════════════════════╗
# ║                  Step 2: 商品详情页截图                           ║
# ╚══════════════════════════════════════════════════════════════════╝

def step2_screenshot(json_path: str, captcha_pause: bool = True):
    """
    读取 JSON 商品列表，启动浏览器逐条打开详情页截图。

    执行流程:
      1. 启动浏览器（自带登录态）
      2. 逐条打开商品详情页，等待渲染 → 截图
      3. 自动跳过已有截图（支持断点续传）
    """
    # --- 2.1 校验依赖 & 路径 ---
    try:
        from playwright.sync_api import sync_playwright  # noqa: F811
    except ImportError:
        log_err("缺少 playwright，请执行: pip install playwright")
        sys.exit(1)

    exe_path = find_chrome_path()
    if not exe_path or not os.path.exists(exe_path):
        log_err(f"Chrome 不存在: {CHROME_EXE}")
        sys.exit(1)

    log_ok(f"浏览器: {exe_path}")
    log_ok(f"用户目录: {CHROME_PROFILE}")

    # --- 2.2 加载数据 ---
    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        items = json.load(f)
    log_step(f"共 {len(items)} 个商品待截图")

    # --- 2.3 增量跳过 ---
    os.makedirs(SCREENSHOTS_DIR, exist_ok=True)
    done_ids = {f.replace(".png", "") for f in os.listdir(SCREENSHOTS_DIR) if f.endswith(".png")}
    pending = [it for it in items if it["itemId"] not in done_ids]
    skipped = len(items) - len(pending)

    if skipped:
        log_ok(f"跳过 {skipped} 个已有截图，待处理 {len(pending)} 个")
    if not pending:
        log_ok("所有商品已有截图，无需处理")
        return

    # --- 2.4 启动浏览器 ---
    success = 0
    fail = 0

    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=CHROME_PROFILE,
                executable_path=exe_path,
                headless=False,
                viewport={"width": 1280, "height": 900},
                args=["--disable-blink-features=AutomationControlled"],
            )
        except Exception:
            log_err("浏览器启动失败！")
            log_err(f"  1. 请确认已关闭使用同一 Profile 的其他 Chrome 窗口")
            log_err(f"  2. 浏览器路径: {exe_path}")
            log_err(f"  3. Profile 目录: {CHROME_PROFILE}")
            log_err(f"\n原始错误: {traceback.format_exc().splitlines()[-1]}")
            sys.exit(1)

        page = context.pages[0] if context.pages else context.new_page()
        log_ok("浏览器已启动，开始截图...")
        print()

        # --- 2.5 逐个截图 ---
        for idx, item in enumerate(pending, 1):
            item_id = item["itemId"]
            detail_url = item.get("detailUrl", "")

            if not detail_url:
                log_warn(f"[{idx}/{len(pending)}] 无详情链接，跳过: {item_id}")
                fail += 1
                continue

            snapshot_path = os.path.join(SCREENSHOTS_DIR, f"{item_id}.png")

            try:
                print(f"  [{idx}/{len(pending)}] {item_id} ... ",
                      end="", flush=True)

                page.goto(detail_url, timeout=30000, wait_until="domcontentloaded")

                # 验证码检测
                if captcha_pause and check_captcha(page):
                    page.wait_for_timeout(2000)

                # 等待页面渲染
                wait_s = random.randint(*PAGE_WAIT)
                page.wait_for_timeout(wait_s * 1000)

                page.screenshot(path=snapshot_path, full_page=False)
                print(f"✓ ({wait_s}s)", flush=True)
                success += 1

                # 商品间停顿
                if idx < len(pending):
                    time.sleep(random.uniform(*BETWEEN_WAIT))
            except Exception:
                print("✗", flush=True)
                log_warn(f"[{idx}/{len(pending)}] {item_id} 失败: "
                         f"{traceback.format_exc().splitlines()[-1]}")
                fail += 1

        context.close()

    # --- 2.6 汇总 ---
    total = success + skipped + fail
    print()
    log_ok(f"截图完成: 新增 {success}, 跳过 {skipped}, 失败 {fail}, 共 {total}/{len(items)}")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 3: JSON → XLSX                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def step3_to_xlsx(json_path: str):
    """
    将搜索结果 JSON + 截图合并为一个 XLSX 文件。
    截图以图片形式嵌入到「商品详情页截图」列。
    """
    # --- 3.1 校验依赖 ---
    try:
        from openpyxl import Workbook  # noqa: F811
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side  # noqa: F811
        from openpyxl.drawing.image import Image as XlImage  # noqa: F811
        from openpyxl.utils import get_column_letter  # noqa: F811
    except ImportError:
        log_err("缺少 openpyxl，请执行: pip install openpyxl Pillow")
        sys.exit(1)

    # --- 3.2 加载数据 ---
    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        items = json.load(f)

    if not items:
        log_err("JSON 文件为空或无数据")
        return

    log_step(f"数据 {len(items)} 条 → 生成 XLSX ...")

    # --- 3.3 写入工作簿 ---
    wb = Workbook()
    ws = wb.active
    ws.title = "闲鱼搜索结果"

    # 表头样式（闲鱼主题色 #FFCC00）
    hdr_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="000000")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 列宽
    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # 数据行
    data_font = Font(name="微软雅黑", size=10)
    screenshot_col = len(HEADERS)  # 最后一列 = 截图列号
    img_ok = 0
    img_miss = 0

    for row_idx, item in enumerate(items, 2):
        row_data = [
            row_idx - 1,
            item.get("seller", ""),
            item.get("title", ""),
            item.get("price", ""),
            item.get("wantCount", ""),
            item.get("detailUrl", ""),
            item.get("image", ""),
            item.get("itemId", ""),
            "",  # 截图列——后面嵌入图片
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            align_h = "center" if col in (1, 4, 5, 8) else "left"
            cell.alignment = Alignment(horizontal=align_h, vertical="center",
                                       wrap_text=(col in (2, 3, 6, 7)))

        # 行高（预留给截图）
        ws.row_dimensions[row_idx].height = IMG_HEIGHT_PT + 8

        # --- 嵌入截图 ---
        item_id = item.get("itemId", "")
        img_path = os.path.join(SCREENSHOTS_DIR, f"{item_id}.png")

        if os.path.exists(img_path):
            try:
                img = XlImage(img_path)
                img.width = IMG_WIDTH_PT
                img.height = IMG_HEIGHT_PT
                ws.add_image(img, f"{get_column_letter(screenshot_col)}{row_idx}")
                img_ok += 1
            except Exception:
                log_warn(f"嵌入截图失败: {item_id}")
                ws.cell(row=row_idx, column=screenshot_col, value="(截图加载失败)")
                img_miss += 1
        else:
            ws.cell(row=row_idx, column=screenshot_col, value="(无截图)")
            img_miss += 1

    # --- 3.4 保存 ---
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(items) + 1}"

    out_path = json_path.rsplit(".", 1)[0] + ".xlsx"
    wb.save(out_path)

    print()
    log_ok(f"XLSX 已保存: {out_path}")
    log_ok(f"数据 {len(items)} 条 | 截图嵌入 {img_ok}/{len(items)} 张")
    if img_miss:
        log_warn(f"{img_miss} 条缺少截图（需先运行 --screenshot）")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                         主入口                                   ║
# ╚══════════════════════════════════════════════════════════════════╝

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    print("=" * 56)
    print("       闲鱼搜索商品采集工具")
    print("=" * 56)

    args = sys.argv[1:]

    # 解析公共可选参数
    pages = DEFAULT_PAGES
    if "--pages" in args:
        pi = args.index("--pages")
        if pi + 1 < len(args):
            try:
                pages = int(args[pi + 1])
            except ValueError:
                log_err("--pages 后面必须是数字")
                sys.exit(1)
    captcha_pause = "--no-captcha" not in args

    if "--auto" in args:
        # 模式3: 一键自动采集（浏览器→提取→截图→XLSX）
        idx = args.index("--auto")
        keyword = args[idx + 1] if idx + 1 < len(args) else ""
        if not keyword or keyword.startswith("--"):
            keyword = input("请输入搜索关键词: ").strip()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)
        step0_auto_collect(keyword, max_pages=pages, captcha_pause=captcha_pause)

    elif "--from-json" in args:
        # 模式2: 根据已有JSON文件：截图+转XLSX
        idx = args.index("--from-json")
        json_file = args[idx + 1] if idx + 1 < len(args) else ""
        if not json_file:
            log_err("请指定 JSON 文件路径")
            sys.exit(1)
        log_step(f"从 JSON 文件开始: {json_file}")
        step2_screenshot(json_file, captcha_pause=captcha_pause)
        print()
        step3_to_xlsx(json_file)

    else:
        # 模式1: 指定关键词，输出提取脚本
        keyword = get_keyword()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)
        step1_generate_js(keyword, max_pages=pages)


if __name__ == "__main__":
    main()
