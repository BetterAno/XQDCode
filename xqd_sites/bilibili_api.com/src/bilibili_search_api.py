"""
B站视频搜索采集工具 - API 协议版
===================================
纯协议请求 + 自动截图 + XLSX 导出，不依赖浏览器自动化采集。

三种运行模式:
  模式1: python src/bilibili_search_api.py 八段锦                              → API搜索，输出JSON
  模式2: python src/bilibili_search_api.py --from-json 八段锦_bilibili_2026-06-17.json  → 根据JSON：截图+转XLSX
  模式3: python src/bilibili_search_api.py --auto 八段锦                       → 一键自动（API搜索→截图→XLSX）

可选参数:
  --pages N         指定采集页数（默认5）
  --no-captcha      关闭验证码暂停检测

依赖: requests, playwright, openpyxl, Pillow
"""

import sys
import io
import os
import json
import hashlib
import datetime
import traceback
import random
import time
import urllib.parse
from functools import reduce

# ╔══════════════════════════════════════════════════════════════════╗
# ║                       0. 配置                                    ║
# ╚══════════════════════════════════════════════════════════════════╝

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SEARCH_API = "https://api.bilibili.com/x/web-interface/wbi/search/type"
NAV_API = "https://api.bilibili.com/x/web-interface/nav"

# 浏览器配置（截图用）
CHROME_EXE = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
CHROME_PROFILE = r"C:\Users\Administrator\Desktop\XQDCode\ChromeDebug"

# 截图延时（秒）
PAGE_WAIT = (1, 2)
BETWEEN_WAIT = (1, 2)

# XLSX 列定义（与 bilibili.com 一致）
HEADERS = [
    "序号", "作者", "标题", "视频ID",
    "播放量", "时长", "发布日期", "状态",
    "详情页链接", "封面图片", "视频详情页截图",
]
COL_WIDTHS = [6, 22, 60, 16, 12, 12, 12, 10, 50, 50, 30]
IMG_WIDTH_PT = 220
IMG_HEIGHT_PT = 155
COVER_WIDTH_PT = 160
COVER_HEIGHT_PT = 120

DEFAULT_PAGES = 5
PAGE_SIZE = 42

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/146.0.0.0 Safari/537.36"
)

# WBI mixin key 索引表
MIXIN_KEY_ENC_TAB = [
    46, 47, 18, 2, 53, 8, 23, 32, 15, 50, 10, 31, 58, 3, 45, 35,
    27, 43, 5, 49, 33, 9, 42, 19, 29, 28, 14, 39, 12, 38, 41, 13,
    37, 48, 7, 16, 24, 55, 40, 61, 26, 17, 0, 1, 60, 51, 30, 4,
    22, 25, 54, 21, 56, 59, 6, 63, 57, 62, 11, 36, 20, 34, 44, 52,
]

# 需要过滤的特殊字符
FILTER_CHARS = set("!*'();:@&=+$,/?%#[]")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                       1. 工具函数                                ║
# ╚══════════════════════════════════════════════════════════════════╝

def _log(level: str, msg: str):
    prefixes = {"INFO": "  ", "OK": "✅", "WARN": "⚠️", "ERR": "❌", "STEP": "📌"}
    print(f"{prefixes.get(level, '  ')} {msg}", flush=True)


def log_step(msg): _log("STEP", msg)
def log_ok(msg): _log("OK", msg)
def log_warn(msg): _log("WARN", msg)
def log_err(msg): _log("ERR", msg)


def find_chrome_path():
    if os.path.exists(CHROME_EXE):
        return CHROME_EXE
    alt = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    return alt if os.path.exists(alt) else None


def make_filename(keyword: str) -> str:
    date_str = datetime.date.today().isoformat()
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return f"{safe_kw}_bilibili_{date_str}.json"


def resolve_json_path(json_path: str) -> str:
    if not json_path.endswith(".json"):
        json_path += ".json"
    if not os.path.isabs(json_path):
        json_path = os.path.join(BASE_DIR, json_path)
    if not os.path.exists(json_path):
        log_err(f"文件不存在: {json_path}")
        sys.exit(1)
    return json_path


def _keyword_from_json(json_path: str) -> str:
    basename = os.path.basename(json_path)
    name = basename.replace(".json", "")
    parts = name.rsplit("_bilibili_", 1)
    return parts[0] if parts else name


def _screenshots_dir(keyword: str) -> str:
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return os.path.join(BASE_DIR, "screenshots", safe_kw)


def check_captcha(page) -> bool:
    """多维度判断B站页面异常"""
    url_lower = page.url.lower()
    reason = ""
    try:
        if "bilibili.com/video/" not in url_lower and "b23.tv" not in url_lower:
            reason = "URL 不是视频详情页"
            raise Exception(reason)

        body_len = page.evaluate("() => (document.body?.innerText || '').length")
        if body_len < 100:
            reason = f"页面文本仅 {body_len} 字符"
            raise Exception(reason)

        title = page.title()
        if any(kw in title for kw in ["验证", "登录", "verify", "login"]):
            reason = f"标题异常: {title}"
            raise Exception(reason)

        return False
    except Exception as e:
        if not reason:
            reason = str(e).replace("Exception: ", "")[:60]

    print()
    print("  ╔══════════════════════════════════════════════╗")
    print(f"  ║  🛑 页面异常 [{reason}]")
    print("  ║     请在浏览器中完成验证/登录                 ║")
    print("  ║     完成后回到此处按 Enter 继续截图            ║")
    print("  ╚══════════════════════════════════════════════╝")
    input("  >>> 按 Enter 继续...")
    print()
    log_ok("已确认，继续截图...")
    return True


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   WBI 签名算法                                    ║
# ╚══════════════════════════════════════════════════════════════════╝

def get_mixin_key(img_key: str, sub_key: str) -> str:
    """从 img_key + sub_key 生成 mixin_key（32字符）"""
    raw = img_key + sub_key
    return reduce(lambda s, i: s + raw[i], MIXIN_KEY_ENC_TAB, '')[:32]


def filter_value(value: str) -> str:
    """过滤特殊字符"""
    return ''.join(c for c in value if c not in FILTER_CHARS)


def sign_wbi(params: dict, mixin_key: str) -> dict:
    """
    WBI 签名
    1. 过滤参数值中的特殊字符
    2. 按 key 排序
    3. 拼接为查询字符串
    4. 附加 wts
    5. MD5(拼接字符串 + mixin_key) → w_rid
    """
    wts = int(time.time())

    # 过滤特殊字符
    filtered = {}
    for k, v in params.items():
        filtered[k] = filter_value(str(v))

    # 添加 wts
    filtered['wts'] = str(wts)

    # 按 key 排序
    sorted_keys = sorted(filtered.keys())
    query_parts = [f"{k}={urllib.parse.quote(filtered[k])}" for k in sorted_keys]
    query_string = '&'.join(query_parts)

    # 计算 w_rid
    w_rid = hashlib.md5((query_string + mixin_key).encode()).hexdigest()

    filtered['w_rid'] = w_rid
    return filtered


def get_wbi_keys(session) -> tuple:
    """从 nav API 获取 WBI 密钥"""
    try:
        resp = session.get(NAV_API, timeout=10)
        data = resp.json()
        wbi_img = data.get('data', {}).get('wbi_img', {})
        img_url = wbi_img.get('img_url', '')
        sub_url = wbi_img.get('sub_url', '')

        img_key = img_url.split('/')[-1].split('.')[0] if img_url else ''
        sub_key = sub_url.split('/')[-1].split('.')[0] if sub_url else ''

        if img_key and sub_key:
            log_ok(f"WBI 密钥: img={img_key[:8]}... sub={sub_key[:8]}...")
            return img_key, sub_key
        else:
            log_err("WBI 密钥为空")
            return '', ''
    except Exception as e:
        log_err(f"获取 WBI 密钥失败: {e}")
        return '', ''


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 1: API 搜索                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def init_session():
    """初始化请求会话（获取 buvid 等基础 Cookie）"""
    import requests

    session = requests.Session()
    session.headers.update({
        "User-Agent": USER_AGENT,
        "Accept": "application/json, text/plain, */*",
        "Origin": "https://search.bilibili.com",
        "Referer": "https://search.bilibili.com/",
    })

    # 访问首页获取基础 Cookie
    try:
        session.get("https://www.bilibili.com/", timeout=10)
        log_ok(f"初始化会话, Cookie: {len(session.cookies)} 个")
    except Exception as e:
        log_warn(f"初始化会话失败: {e}")

    return session


def search_once(keyword, page_num, mixin_key, session):
    """执行一次搜索请求"""
    params = {
        "category_id": "",
        "search_type": "video",
        "ad_resource": "5654",
        "__refresh__": "true",
        "_extra": "",
        "context": "",
        "page": str(page_num),
        "page_size": str(PAGE_SIZE),
        "pubtime_begin_s": "0",
        "pubtime_end_s": "0",
        "from_spmid": "333.337",
        "platform": "pc",
        "highlight": "1",
        "single_column": "0",
        "keyword": keyword,
        "source_tag": "3",
        "gaia_vtoken": "",
        "dynamic_offset": str((page_num - 1) * PAGE_SIZE),
        "web_roll_page": "1",
        "web_location": "1430654",
    }

    # WBI 签名
    signed_params = sign_wbi(params, mixin_key)

    try:
        resp = session.get(SEARCH_API, params=signed_params, timeout=15)
        data = resp.json()

        if data.get("code") != 0:
            log_err(f"API 错误 {data.get('code')}: {data.get('message', '')}")
            return []

        result = data.get("data", {}).get("result", [])
        items = []
        for item in result:
            if item.get("type") != "video":
                continue

            # 清理标题中的 HTML 标签
            title = item.get("title", "")
            title = title.replace('<em class="keyword">', '').replace('</em>', '')

            # 封面图补全协议
            pic = item.get("pic", "")
            if pic.startswith("//"):
                pic = "https:" + pic

            # 发布日期
            pubdate = item.get("pubdate", 0)
            pub_date = datetime.datetime.fromtimestamp(pubdate).strftime("%Y-%m-%d") if pubdate else ""

            items.append({
                "videoId": item.get("bvid", ""),
                "title": title,
                "author": item.get("author", ""),
                "publishDate": pub_date,
                "detailUrl": f"https://www.bilibili.com/video/{item.get('bvid', '')}/",
                "image": pic,
                "playCount": str(item.get("play", 0)),
                "danmaku": str(item.get("danmaku", 0)),
                "duration": item.get("duration", ""),
                "aid": item.get("aid", 0),
                "mid": item.get("mid", 0),
                "tag": item.get("tag", ""),
                "description": item.get("description", ""),
                "favorites": item.get("favorites", 0),
                "like": item.get("like", 0),
                "review": item.get("review", 0),
            })

        return items
    except Exception as e:
        log_err(f"请求异常: {e}")
        return []


def step1_api_search(keyword, max_pages=DEFAULT_PAGES):
    """API 搜索多页，返回 JSON 路径"""
    log_step(f"API 搜索: {keyword} | 页数: {max_pages}")

    session = init_session()
    img_key, sub_key = get_wbi_keys(session)
    if not img_key or not sub_key:
        log_err("无法获取 WBI 密钥")
        return None

    mixin_key = get_mixin_key(img_key, sub_key)
    log_ok(f"mixin_key: {mixin_key[:16]}...")

    all_items = []
    seen_ids = set()

    for page_num in range(1, max_pages + 1):
        log_step(f"第 {page_num} 页")

        items = search_once(keyword, page_num, mixin_key, session)
        if not items:
            log_warn(f"第 {page_num} 页无数据，停止")
            break

        new_items = [i for i in items if i["videoId"] not in seen_ids]
        for i in new_items:
            seen_ids.add(i["videoId"])
        all_items.extend(new_items)
        log_ok(f"第 {page_num} 页: {len(new_items)} 条新 (总计 {len(all_items)})")

        if page_num < max_pages:
            delay = random.uniform(1.0, 2.0)
            time.sleep(delay)

    if not all_items:
        log_err("未获取到数据")
        return None

    filename = make_filename(keyword)
    json_path = os.path.join(BASE_DIR, filename)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_items, f, ensure_ascii=False, indent=2)
    log_ok(f"保存 {len(all_items)} 条 → {json_path}")
    return json_path


# ╔══════════════════════════════════════════════════════════════════╗
# ║                  Step 2: 视频详情页截图                           ║
# ╚══════════════════════════════════════════════════════════════════╝

def step2_screenshot(json_path: str, captcha_pause: bool = True, keyword: str = None):
    """读取 JSON，启动浏览器逐条截图"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log_err("缺少 playwright，请执行: pip install playwright")
        sys.exit(1)

    exe_path = find_chrome_path()
    if not exe_path or not os.path.exists(exe_path):
        log_err(f"Chrome 不存在: {CHROME_EXE}")
        sys.exit(1)

    log_ok(f"浏览器: {exe_path}")

    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = raw if isinstance(raw, list) else raw.get("data", [])
    log_step(f"共 {len(items)} 个视频待截图")

    if not keyword:
        keyword = _keyword_from_json(json_path)
    screenshots_dir = _screenshots_dir(keyword)
    os.makedirs(screenshots_dir, exist_ok=True)
    done_ids = {f.replace(".png", "") for f in os.listdir(screenshots_dir) if f.endswith(".png")}
    pending = [it for it in items if it.get("videoId", "") not in done_ids]
    skipped = len(items) - len(pending)

    if skipped:
        log_ok(f"跳过 {skipped} 个已有截图，待处理 {len(pending)} 个")
    if not pending:
        log_ok("所有视频已有截图")
        return done_ids

    success = fail = 0
    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=CHROME_PROFILE, executable_path=exe_path,
                headless=False, viewport={"width": 1280, "height": 900},
                args=["--disable-blink-features=AutomationControlled"])
        except Exception:
            log_err("浏览器启动失败！请关闭占用同 Profile 的 Chrome 窗口")
            sys.exit(1)

        page = context.pages[0] if context.pages else context.new_page()
        log_ok("浏览器已启动，开始截图...")

        for idx, item in enumerate(pending, 1):
            video_id = item.get("videoId", "")
            detail_url = item.get("detailUrl", "")
            if not detail_url:
                log_warn(f"[{idx}/{len(pending)}] 无详情链接，跳过: {video_id}")
                fail += 1
                continue

            snapshot_path = os.path.join(screenshots_dir, f"{video_id}.png")
            try:
                print(f"  [{idx}/{len(pending)}] {video_id} ... ", end="", flush=True)
                page.goto(detail_url, timeout=30000, wait_until="domcontentloaded")
                if captcha_pause and check_captcha(page):
                    page.wait_for_timeout(2000)
                wait_s = random.randint(*PAGE_WAIT)
                page.wait_for_timeout(wait_s * 1000)
                page.screenshot(path=snapshot_path, full_page=False)
                print(f"✓ ({wait_s}s)", flush=True)
                success += 1
                if idx < len(pending):
                    time.sleep(random.uniform(*BETWEEN_WAIT))
            except Exception:
                print("✗", flush=True)
                log_warn(f"[{idx}/{len(pending)}] {video_id} 失败")
                fail += 1

        context.close()

    total = success + skipped + fail
    print()
    log_ok(f"截图完成: 新增 {success}, 跳过 {skipped}, 失败 {fail}, 共 {total}/{len(items)}")
    return done_ids


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 3: JSON → XLSX                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def step3_to_xlsx(json_path: str, existing_ids: set = None, keyword: str = None):
    """将 JSON + 截图合并为 XLSX"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
    except ImportError:
        log_err("缺少 openpyxl，请执行: pip install openpyxl Pillow")
        sys.exit(1)

    import urllib.request
    import io as _io

    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = raw if isinstance(raw, list) else raw.get("data", [])
    if not items:
        log_err("JSON 文件为空")
        return

    log_step(f"数据 {len(items)} 条 → 生成 XLSX ...")

    if not keyword:
        keyword = _keyword_from_json(json_path)
    screenshots_dir = _screenshots_dir(keyword)

    if existing_ids is None:
        existing_ids = set()
        if os.path.exists(screenshots_dir):
            existing_ids = {f.replace(".png", "") for f in os.listdir(screenshots_dir) if f.endswith(".png")}

    wb = Workbook()
    ws = wb.active
    ws.title = "B站搜索结果"

    # 表头样式（B站粉色）
    hdr_fill = PatternFill(start_color="00A1D6", end_color="00A1D6", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    data_font = Font(name="微软雅黑", size=10)
    screenshot_col = len(HEADERS)
    cover_col = 10  # 封面图片列
    img_ok = img_miss = cover_ok = cover_miss = 0

    for row_idx, item in enumerate(items, 2):
        video_id = item.get("videoId", "")
        status = "已爬取" if video_id in existing_ids else "新增"
        row_data = [
            row_idx - 1,
            item.get("author", ""),
            item.get("title", ""),
            video_id,
            item.get("playCount", ""),
            item.get("duration", ""),
            item.get("publishDate", ""),
            status,
            item.get("detailUrl", ""),
            "",  # 封面图
            "",  # 截图
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            align_h = "center" if col in (1, 4, 5, 6, 7, 8) else "left"
            cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=(col in (2, 3, 9)))

        ws.row_dimensions[row_idx].height = max(IMG_HEIGHT_PT, COVER_HEIGHT_PT) + 8

        # 封面图
        cover_url = item.get("image", "")
        if cover_url:
            try:
                if cover_url.startswith("//"):
                    cover_url = "https:" + cover_url
                req = urllib.request.Request(cover_url, headers={
                    "User-Agent": USER_AGENT, "Referer": "https://www.bilibili.com/"})
                with urllib.request.urlopen(req, timeout=15) as resp:
                    img_data = resp.read()
                cover_img = XlImage(_io.BytesIO(img_data))
                cover_img.width = COVER_WIDTH_PT
                cover_img.height = COVER_HEIGHT_PT
                ws.add_image(cover_img, f"{get_column_letter(cover_col)}{row_idx}")
                cover_ok += 1
            except Exception:
                ws.cell(row=row_idx, column=cover_col, value=cover_url)
                cover_miss += 1
        else:
            cover_miss += 1

        # 截图
        img_path = os.path.join(screenshots_dir, f"{video_id}.png")
        if os.path.exists(img_path):
            try:
                img = XlImage(img_path)
                img.width = IMG_WIDTH_PT
                img.height = IMG_HEIGHT_PT
                ws.add_image(img, f"{get_column_letter(screenshot_col)}{row_idx}")
                img_ok += 1
            except Exception:
                ws.cell(row=row_idx, column=screenshot_col, value="(截图加载失败)")
                img_miss += 1
        else:
            ws.cell(row=row_idx, column=screenshot_col, value="(无截图)")
            img_miss += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(items) + 1}"

    out_path = json_path.rsplit(".", 1)[0] + ".xlsx"
    wb.save(out_path)

    print()
    log_ok(f"XLSX 已保存: {out_path}")
    log_ok(f"数据 {len(items)} 条 | 封面图 {cover_ok}/{len(items)} | 截图 {img_ok}/{len(items)}")
    if cover_miss:
        log_warn(f"{cover_miss} 条封面图下载失败")
    if img_miss:
        log_warn(f"{img_miss} 条缺少截图")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                         主入口                                   ║
# ╚══════════════════════════════════════════════════════════════════╝

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    print("=" * 56)
    print("    B站视频搜索采集工具 - API 协议版")
    print("=" * 56)

    args = sys.argv[1:]

    pages = DEFAULT_PAGES
    if "--pages" in args:
        pi = args.index("--pages")
        if pi + 1 < len(args):
            pages = int(args[pi + 1])
    captcha_pause = "--no-captcha" not in args

    if "--auto" in args:
        # 模式3: API搜索 → 截图 → XLSX
        idx = args.index("--auto")
        keyword = args[idx + 1] if idx + 1 < len(args) else ""
        if not keyword or keyword.startswith("--"):
            keyword = input("请输入搜索关键词: ").strip()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)

        json_path = step1_api_search(keyword, max_pages=pages)
        if not json_path:
            sys.exit(1)
        print()
        log_step("继续截图...")
        existing_ids = step2_screenshot(json_path, captcha_pause=captcha_pause, keyword=keyword)
        print()
        log_step("继续生成 XLSX...")
        step3_to_xlsx(json_path, existing_ids=existing_ids, keyword=keyword)

    elif "--from-json" in args:
        # 模式2: JSON → 截图 → XLSX
        idx = args.index("--from-json")
        json_file = args[idx + 1] if idx + 1 < len(args) else ""
        if not json_file:
            log_err("请指定 JSON 文件路径")
            sys.exit(1)
        keyword = _keyword_from_json(json_file)
        existing_ids = step2_screenshot(json_file, captcha_pause=captcha_pause, keyword=keyword)
        print()
        step3_to_xlsx(json_file, existing_ids=existing_ids, keyword=keyword)

    else:
        # 模式1: API 搜索 → JSON
        keyword = args[0] if args else ""
        if not keyword or keyword.startswith("--"):
            keyword = input("请输入搜索关键词: ").strip()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)
        step1_api_search(keyword, max_pages=pages)


if __name__ == "__main__":
    main()
