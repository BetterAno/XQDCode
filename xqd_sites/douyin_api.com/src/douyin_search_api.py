"""
抖音视频搜索采集工具 - API 协议版
===================================
三种运行模式:
  模式1: python src/douyin_search_api.py 八段锦                              → API搜索，输出JSON
  模式2: python src/douyin_search_api.py --from-json 八段锦_douyin_2026-06-17.json  → 根据JSON：截图+转XLSX
  模式3: python src/douyin_search_api.py --auto 八段锦                       → 一键自动（API搜索→截图→XLSX）

可选参数:
  --pages N         指定采集页数（默认5）
  --no-captcha      关闭验证码暂停检测

依赖: requests, node (with webmssdk), playwright, openpyxl, Pillow
"""

import sys
import io
import os
import json
import hashlib
import subprocess
import datetime
import traceback
import random
import time
import urllib.parse

# ╔══════════════════════════════════════════════════════════════════╗
# ║                       0. 配置                                    ║
# ╚══════════════════════════════════════════════════════════════════╝

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SIGNER_JS = os.path.join(BASE_DIR, "src", "signer.js")
COOKIES_FILE = os.path.join(BASE_DIR, "cookies.txt")
NODE_EXE = "node"

SEARCH_API = "https://www.douyin.com/aweme/v1/web/general/search/single/"

# 浏览器配置（截图用）
CHROME_EXE = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
CHROME_PROFILE = r"C:\Users\Administrator\Desktop\XQDCode\ChromeDebug"

# 截图延时（秒）
PAGE_WAIT = (2, 3)
BETWEEN_WAIT = (3, 4)

# XLSX 列定义
HEADERS = [
    "序号", "作者", "标题", "视频ID",
    "状态", "详情页链接", "视频图片", "视频详情页截图",
]
COL_WIDTHS = [6, 22, 60, 20, 10, 50, 50, 30]
IMG_WIDTH_PT = 220
IMG_HEIGHT_PT = 155
COVER_WIDTH_PT = 160
COVER_HEIGHT_PT = 210

DEFAULT_PAGES = 5
COUNT_PER_PAGE = 10

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/146.0.0.0 Safari/537.36"
)

# API 固定参数
FIXED_PARAMS = {
    "device_platform": "webapp", "aid": "6383", "channel": "channel_pc_web",
    "search_channel": "aweme_general", "enable_history": "1",
    "search_source": "normal_search", "query_correct_type": "1",
    "is_filter_search": "0", "from_group_id": "", "disable_rs": "0",
    "need_filter_settings": "0", "list_type": "single", "pc_client_type": "1",
    "version_code": "190600", "version_name": "19.6.0", "cookie_enabled": "true",
    "screen_width": "1920", "screen_height": "1080", "browser_language": "zh-CN",
    "browser_platform": "Win32", "browser_name": "Chrome",
    "browser_version": "146.0.0.0", "browser_online": "true",
    "engine_name": "Blink", "engine_version": "146.0.0.0",
    "os_name": "Windows", "os_version": "10", "device_memory": "8",
    "platform": "PC", "downlink": "10", "effective_type": "4g",
    "round_trip_time": "0",
}

# 签名参数列表（参与 X-MS-STUB 计算）
SIGN_PARAMS = [
    "aid", "channel", "search_channel", "keyword",
    "search_source", "offset", "count", "list_type",
    "device_platform", "version_code", "version_name",
]


# ╔══════════════════════════════════════════════════════════════════╗
# ║                       1. 工具函数                                ║
# ╚══════════════════════════════════════════════════════════════════╝

def _log(level: str, msg: str):
    prefixes = {"INFO": "  ", "OK": "✅", "WARN": "⚠️", "ERR": "❌", "STEP": "📌"}
    print(f"{prefixes.get(level, '  ')} {msg}", flush=True)


def log_step(msg): _log("STEP", msg)
def log_ok(msg): _log("OK", msg)
def log_warn(msg): _log("WARN", msg)
def log_err(msg): _log("ERR", msg)


def find_chrome_path():
    if os.path.exists(CHROME_EXE):
        return CHROME_EXE
    alt = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    return alt if os.path.exists(alt) else None


def make_filename(keyword: str) -> str:
    date_str = datetime.date.today().isoformat()
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return f"{safe_kw}_douyin_{date_str}.json"


def resolve_json_path(json_path: str) -> str:
    if not json_path.endswith(".json"):
        json_path += ".json"
    if not os.path.isabs(json_path):
        json_path = os.path.join(BASE_DIR, json_path)
    if not os.path.exists(json_path):
        log_err(f"文件不存在: {json_path}")
        sys.exit(1)
    return json_path


def _keyword_from_json(json_path: str) -> str:
    basename = os.path.basename(json_path)
    name = basename.replace(".json", "")
    parts = name.rsplit("_douyin_", 1)
    return parts[0] if parts else name


def _screenshots_dir(keyword: str) -> str:
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return os.path.join(BASE_DIR, "screenshots", safe_kw)


def check_captcha(page) -> bool:
    """多维度判断抖音页面异常（验证码/登录/风控）"""
    url_lower = page.url.lower()
    reason = ""
    try:
        try:
            page.wait_for_selector(
                ".captcha_verify_container, .verify-captcha, .login-modal, "
                ".douyin-login, #secsdk-captcha-drag-wrapper, .captcha-box",
                timeout=5000, state="attached")
        except Exception:
            pass

        if "douyin.com/video/" not in url_lower:
            reason = "URL 不是视频详情页"
            raise Exception(reason)

        if any(kw in url_lower for kw in ["login.", "verify.", "captcha", "punish", "sec-verify"]):
            reason = "URL 包含验证关键词"
            raise Exception(reason)

        body_len = page.evaluate("() => (document.body?.innerText || '').length")
        if body_len < 100:
            reason = f"页面文本仅 {body_len} 字符"
            raise Exception(reason)

        has_video = page.evaluate(r"""() => {
          const t = (document.body?.innerText || '').slice(0, 2000).toLowerCase();
          const signals = ['点赞','评论','收藏','分享','关注','粉丝','作品'];
          let n = 0; for (const s of signals) { if (t.includes(s)) n++; }
          return n >= 2;
        }""")
        if not has_video:
            reason = "页面缺少视频特征"
            raise Exception(reason)

        title = page.title()
        if any(kw in title for kw in ["验证", "登录", "verify", "login", "安全", "滑块"]):
            reason = f"标题异常: {title}"
            raise Exception(reason)

        return False
    except Exception as e:
        if not reason:
            reason = str(e).replace("Exception: ", "")[:60]

    print()
    print("  ╔══════════════════════════════════════════════╗")
    print(f"  ║  🛑 页面异常 [{reason}]")
    print("  ║     请在浏览器中完成验证/登录                 ║")
    print("  ║     完成后回到此处按 Enter 继续截图            ║")
    print("  ╚══════════════════════════════════════════════╝")
    input("  >>> 按 Enter 继续...")
    print()
    log_ok("已确认，继续截图...")
    return True


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 1: API 搜索                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def md5_hash(text):
    return hashlib.md5(text.encode('utf-8')).hexdigest()


def generate_stub(params):
    parts = [f"{name}={params.get(name, '')}" for name in SIGN_PARAMS]
    return md5_hash(",".join(parts))


def call_signer(stub_hash):
    try:
        result = subprocess.run(
            [NODE_EXE, SIGNER_JS, stub_hash],
            capture_output=True, text=True, timeout=15, cwd=BASE_DIR)
        if result.returncode != 0:
            log_err(f"签名器错误: {result.stderr[:200]}")
            return None
        return result.stdout.strip()
    except Exception as e:
        log_err(f"签名器异常: {e}")
        return None


def generate_search_id():
    ts = int(time.time() * 1000)
    rand = random.randbytes(8).hex().upper()
    return f"{ts}{rand}"[:32]


def load_cookies():
    if not os.path.exists(COOKIES_FILE):
        log_err(f"Cookie 文件不存在: {COOKIES_FILE}")
        log_info("请从浏览器复制完整 Cookie 到 cookies.txt")
        return {}
    with open(COOKIES_FILE, "r", encoding="utf-8") as f:
        cookie_str = f.read().strip()
    cookies = {}
    for item in cookie_str.split(";"):
        item = item.strip()
        if "=" in item:
            k, v = item.split("=", 1)
            cookies[k.strip()] = v.strip()
    log_ok(f"加载 Cookie: {len(cookies)} 个")
    return cookies


def search_once(keyword, offset, cookies, session):
    """执行一次搜索请求"""
    params = dict(FIXED_PARAMS)
    params["keyword"] = keyword
    params["offset"] = str(offset)
    params["count"] = str(COUNT_PER_PAGE)
    params["search_id"] = generate_search_id()
    params["pc_search_top_1_params"] = '{"enable_ai_search_top_1":1}'

    fp = cookies.get("s_v_web_id", "verify_mqhsqvpt_m8AYka4O_TNFY_4MeF_BLoI_M7tEHWs5H8yB")
    params["verifyFp"] = fp
    params["fp"] = fp
    if "UIFID" in cookies:
        params["uifid"] = cookies["UIFID"]
    if "msToken" in cookies:
        params["msToken"] = cookies["msToken"]

    stub = generate_stub(params)
    x_bogus = call_signer(stub)
    if not x_bogus:
        log_err("签名失败")
        return [], cookies
    params["a_bogus"] = x_bogus

    cookie_str = "; ".join(f"{k}={v}" for k, v in cookies.items())
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "application/json, text/plain, */*",
        "Referer": f"https://www.douyin.com/search/{urllib.parse.quote(keyword)}",
        "Cookie": cookie_str,
        "sec-ch-ua": '"Chromium";v="146", "Not-A.Brand";v="24", "Google Chrome";v="146"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    url = SEARCH_API + "?" + urllib.parse.urlencode(params)
    try:
        resp = session.get(url, headers=headers, timeout=15)
        for k, v in resp.cookies.items():
            cookies[k] = v

        if resp.status_code != 200:
            log_err(f"HTTP {resp.status_code}")
            return [], cookies

        data = resp.json()
        if data.get("status_code") != 0:
            log_err(f"API 错误 {data.get('status_code')}: {data.get('status_msg', '')}")
            return [], cookies

        items = []
        for item in data.get("data", []):
            aweme = item.get("aweme_info", {})
            if not aweme:
                continue
            author = aweme.get("author", {})
            video = aweme.get("video", {})
            cover = video.get("cover", {})
            cover_urls = cover.get("url_list", [])
            items.append({
                "videoId": aweme.get("aweme_id", ""),
                "title": (aweme.get("desc", "") or "").replace("\n", " "),
                "author": author.get("nickname", ""),
                "detailUrl": f"https://www.douyin.com/video/{aweme.get('aweme_id', '')}",
                "image": cover_urls[0] if cover_urls else "",
                "createTime": aweme.get("create_time", 0),
                "uid": author.get("uid", ""),
                "secUid": author.get("sec_uid", ""),
            })
        return items, cookies
    except Exception as e:
        log_err(f"请求异常: {e}")
        return [], cookies


def step1_api_search(keyword, max_pages=DEFAULT_PAGES):
    """API 搜索多页，返回 JSON 路径"""
    import requests

    log_step(f"API 搜索: {keyword} | 页数: {max_pages}")

    cookies = load_cookies()
    if not cookies:
        return None

    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})

    all_items = []
    seen_ids = set()

    for page in range(max_pages):
        offset = page * COUNT_PER_PAGE
        log_step(f"第 {page + 1} 页 | offset={offset}")

        items, cookies = search_once(keyword, offset, cookies, session)
        if not items:
            log_warn(f"第 {page + 1} 页无数据，停止")
            break

        new_items = [i for i in items if i["videoId"] not in seen_ids]
        for i in new_items:
            seen_ids.add(i["videoId"])
        all_items.extend(new_items)
        log_ok(f"第 {page + 1} 页: {len(new_items)} 条新 (总计 {len(all_items)})")

        if page < max_pages - 1:
            delay = random.uniform(1.5, 3.0)
            time.sleep(delay)

    if not all_items:
        log_err("未获取到数据")
        return None

    filename = make_filename(keyword)
    json_path = os.path.join(BASE_DIR, filename)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_items, f, ensure_ascii=False, indent=2)
    log_ok(f"保存 {len(all_items)} 条 → {json_path}")
    return json_path


# ╔══════════════════════════════════════════════════════════════════╗
# ║                  Step 2: 视频详情页截图                           ║
# ╚══════════════════════════════════════════════════════════════════╝

def step2_screenshot(json_path: str, captcha_pause: bool = True, keyword: str = None):
    """读取 JSON，启动浏览器逐条截图（增量跳过已有截图）"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log_err("缺少 playwright，请执行: pip install playwright")
        sys.exit(1)

    exe_path = find_chrome_path()
    if not exe_path or not os.path.exists(exe_path):
        log_err(f"Chrome 不存在: {CHROME_EXE}")
        sys.exit(1)

    log_ok(f"浏览器: {exe_path}")

    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = raw if isinstance(raw, list) else raw.get("data", [])
    log_step(f"共 {len(items)} 个视频待截图")

    if not keyword:
        keyword = _keyword_from_json(json_path)
    screenshots_dir = _screenshots_dir(keyword)
    os.makedirs(screenshots_dir, exist_ok=True)
    done_ids = {f.replace(".png", "") for f in os.listdir(screenshots_dir) if f.endswith(".png")}
    pending = [it for it in items if it.get("videoId", "") not in done_ids]
    skipped = len(items) - len(pending)

    if skipped:
        log_ok(f"跳过 {skipped} 个已有截图，待处理 {len(pending)} 个")
    if not pending:
        log_ok("所有视频已有截图")
        return done_ids

    success = fail = 0
    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=CHROME_PROFILE, executable_path=exe_path,
                headless=False, viewport={"width": 1280, "height": 900},
                args=["--disable-blink-features=AutomationControlled"])
        except Exception:
            log_err("浏览器启动失败！请关闭占用同 Profile 的 Chrome 窗口")
            sys.exit(1)

        page = context.pages[0] if context.pages else context.new_page()
        log_ok("浏览器已启动，开始截图...")

        for idx, item in enumerate(pending, 1):
            video_id = item.get("videoId", "")
            detail_url = item.get("detailUrl", "")
            if not detail_url:
                log_warn(f"[{idx}/{len(pending)}] 无详情链接，跳过: {video_id}")
                fail += 1
                continue

            snapshot_path = os.path.join(screenshots_dir, f"{video_id}.png")
            try:
                print(f"  [{idx}/{len(pending)}] {video_id} ... ", end="", flush=True)
                page.goto(detail_url, timeout=30000, wait_until="domcontentloaded")
                if captcha_pause and check_captcha(page):
                    page.wait_for_timeout(2000)
                wait_s = random.randint(*PAGE_WAIT)
                page.wait_for_timeout(wait_s * 1000)
                page.screenshot(path=snapshot_path, full_page=False)
                print(f"✓ ({wait_s}s)", flush=True)
                success += 1
                if idx < len(pending):
                    time.sleep(random.uniform(*BETWEEN_WAIT))
            except Exception:
                print("✗", flush=True)
                log_warn(f"[{idx}/{len(pending)}] {video_id} 失败")
                fail += 1

        context.close()

    total = success + skipped + fail
    print()
    log_ok(f"截图完成: 新增 {success}, 跳过 {skipped}, 失败 {fail}, 共 {total}/{len(items)}")
    return done_ids


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 3: JSON → XLSX                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def step3_to_xlsx(json_path: str, existing_ids: set = None, keyword: str = None):
    """将 JSON + 截图合并为 XLSX"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
    except ImportError:
        log_err("缺少 openpyxl，请执行: pip install openpyxl Pillow")
        sys.exit(1)

    import urllib.request
    import io as _io

    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = raw if isinstance(raw, list) else raw.get("data", [])
    if not items:
        log_err("JSON 文件为空")
        return

    log_step(f"数据 {len(items)} 条 → 生成 XLSX ...")

    if not keyword:
        keyword = _keyword_from_json(json_path)
    screenshots_dir = _screenshots_dir(keyword)

    if existing_ids is None:
        existing_ids = set()
        if os.path.exists(screenshots_dir):
            existing_ids = {f.replace(".png", "") for f in os.listdir(screenshots_dir) if f.endswith(".png")}

    wb = Workbook()
    ws = wb.active
    ws.title = "抖音搜索结果"

    hdr_fill = PatternFill(start_color="FF0050", end_color="FF0050", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    data_font = Font(name="微软雅黑", size=10)
    screenshot_col = len(HEADERS)
    cover_col = 7
    img_ok = img_miss = cover_ok = cover_miss = 0

    for row_idx, item in enumerate(items, 2):
        video_id = item.get("videoId", "")
        status = "已爬取" if video_id in existing_ids else "新增"
        row_data = [
            row_idx - 1, item.get("author", ""), item.get("title", ""),
            video_id, status, item.get("detailUrl", ""), "", ""]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            align_h = "center" if col in (1, 4, 5) else "left"
            cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=(col in (2, 3, 6)))

        ws.row_dimensions[row_idx].height = max(IMG_HEIGHT_PT, COVER_HEIGHT_PT) + 8

        # 封面图
        cover_url = item.get("image", "")
        if cover_url:
            try:
                req = urllib.request.Request(cover_url, headers={
                    "User-Agent": USER_AGENT, "Referer": "https://www.douyin.com/"})
                with urllib.request.urlopen(req, timeout=15) as resp:
                    img_data = resp.read()
                cover_img = XlImage(_io.BytesIO(img_data))
                cover_img.width = COVER_WIDTH_PT
                cover_img.height = COVER_HEIGHT_PT
                ws.add_image(cover_img, f"{get_column_letter(cover_col)}{row_idx}")
                cover_ok += 1
            except Exception:
                ws.cell(row=row_idx, column=cover_col, value=cover_url)
                cover_miss += 1
        else:
            cover_miss += 1

        # 截图
        img_path = os.path.join(screenshots_dir, f"{video_id}.png")
        if os.path.exists(img_path):
            try:
                img = XlImage(img_path)
                img.width = IMG_WIDTH_PT
                img.height = IMG_HEIGHT_PT
                ws.add_image(img, f"{get_column_letter(screenshot_col)}{row_idx}")
                img_ok += 1
            except Exception:
                ws.cell(row=row_idx, column=screenshot_col, value="(截图加载失败)")
                img_miss += 1
        else:
            ws.cell(row=row_idx, column=screenshot_col, value="(无截图)")
            img_miss += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(items) + 1}"

    out_path = json_path.rsplit(".", 1)[0] + ".xlsx"
    wb.save(out_path)

    print()
    log_ok(f"XLSX 已保存: {out_path}")
    log_ok(f"数据 {len(items)} 条 | 封面图 {cover_ok}/{len(items)} | 截图 {img_ok}/{len(items)}")
    if cover_miss:
        log_warn(f"{cover_miss} 条封面图下载失败")
    if img_miss:
        log_warn(f"{img_miss} 条缺少截图")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                         主入口                                   ║
# ╚══════════════════════════════════════════════════════════════════╝

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    print("=" * 56)
    print("    抖音视频搜索采集工具 - API 协议版")
    print("=" * 56)

    args = sys.argv[1:]

    pages = DEFAULT_PAGES
    if "--pages" in args:
        pi = args.index("--pages")
        if pi + 1 < len(args):
            pages = int(args[pi + 1])
    captcha_pause = "--no-captcha" not in args

    if "--auto" in args:
        # 模式3: API搜索 → 截图 → XLSX
        idx = args.index("--auto")
        keyword = args[idx + 1] if idx + 1 < len(args) else ""
        if not keyword or keyword.startswith("--"):
            keyword = input("请输入搜索关键词: ").strip()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)

        json_path = step1_api_search(keyword, max_pages=pages)
        if not json_path:
            sys.exit(1)
        print()
        log_step("继续截图...")
        existing_ids = step2_screenshot(json_path, captcha_pause=captcha_pause, keyword=keyword)
        print()
        log_step("继续生成 XLSX...")
        step3_to_xlsx(json_path, existing_ids=existing_ids, keyword=keyword)

    elif "--from-json" in args:
        # 模式2: JSON → 截图 → XLSX
        idx = args.index("--from-json")
        json_file = args[idx + 1] if idx + 1 < len(args) else ""
        if not json_file:
            log_err("请指定 JSON 文件路径")
            sys.exit(1)
        keyword = _keyword_from_json(json_file)
        existing_ids = step2_screenshot(json_file, captcha_pause=captcha_pause, keyword=keyword)
        print()
        step3_to_xlsx(json_file, existing_ids=existing_ids, keyword=keyword)

    else:
        # 模式1: API 搜索 → JSON
        keyword = args[0] if args else ""
        if not keyword or keyword.startswith("--"):
            keyword = input("请输入搜索关键词: ").strip()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)
        step1_api_search(keyword, max_pages=pages)


if __name__ == "__main__":
    main()
