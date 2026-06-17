"""
闲鱼商品搜索采集工具 - API 协议版
===================================
纯协议请求 + 自动截图 + XLSX 导出，不依赖浏览器自动化采集。

三种运行模式:
  模式1: python src/xianyu_search_api.py 手机壳                              → API搜索，输出JSON
  模式2: python src/xianyu_search_api.py --from-json 手机壳_xianyu_2026-06-17.json  → 根据JSON：截图+转XLSX
  模式3: python src/xianyu_search_api.py --auto 手机壳                       → 一键自动（API搜索→截图→XLSX）

可选参数:
  --pages N         指定采集页数（默认5）
  --no-captcha      关闭验证码暂停检测

依赖: requests, playwright, openpyxl, Pillow

注意:
  闲鱼搜索需要登录态 Cookie（_m_h5_tk 等）。
  签名算法: sign = md5(token + '&' + t + '&' + appKey + '&' + data)
"""

import sys
import io
import os
import json
import hashlib
import datetime
import random
import time
import urllib.parse

# ╔══════════════════════════════════════════════════════════════════╗
# ║                       0. 配置                                    ║
# ╚══════════════════════════════════════════════════════════════════╝

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
COOKIES_FILE = os.path.join(BASE_DIR, "cookies.txt")

SEARCH_API = "https://h5api.m.goofish.com/h5/mtop.taobao.idlemtopsearch.pc.search/1.0/"
APP_KEY = "34839810"

# 浏览器配置（截图用）
CHROME_EXE = r"C:\Users\Administrator\AppData\Local\Google\Chrome\Application\chrome.exe"
CHROME_PROFILE = r"C:\Users\Administrator\Desktop\XQDCode\ChromeDebug"

# 截图延时（秒）
PAGE_WAIT = (2, 3)
BETWEEN_WAIT = (2, 3)

# XLSX 列定义（与 xianyu.com 一致）
HEADERS = [
    "序号", "标题", "商品ID", "价格",
    "想要人数", "卖家名", "地区", "状态", "详情页链接",
    "商品图片", "商品详情页截图",
]
COL_WIDTHS = [6, 60, 18, 10, 12, 18, 10, 10, 55, 50, 30]
IMG_WIDTH_PT = 220
IMG_HEIGHT_PT = 155
COVER_WIDTH_PT = 160
COVER_HEIGHT_PT = 160

DEFAULT_PAGES = 5
ROWS_PER_PAGE = 30

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/146.0.0.0 Safari/537.36"
)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                       1. 工具函数                                ║
# ╚══════════════════════════════════════════════════════════════════╝

def _log(level: str, msg: str):
    prefixes = {"INFO": "  ", "OK": "✅", "WARN": "⚠️", "ERR": "❌", "STEP": "📌"}
    print(f"{prefixes.get(level, '  ')} {msg}", flush=True)


def log_step(msg): _log("STEP", msg)
def log_ok(msg): _log("OK", msg)
def log_warn(msg): _log("WARN", msg)
def log_err(msg): _log("ERR", msg)


def find_chrome_path():
    if os.path.exists(CHROME_EXE):
        return CHROME_EXE
    alt = r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    return alt if os.path.exists(alt) else None


def make_filename(keyword: str) -> str:
    date_str = datetime.date.today().isoformat()
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return f"{safe_kw}_xianyu_{date_str}.json"


def resolve_json_path(json_path: str) -> str:
    if not json_path.endswith(".json"):
        json_path += ".json"
    if not os.path.isabs(json_path):
        json_path = os.path.join(BASE_DIR, json_path)
    if not os.path.exists(json_path):
        log_err(f"文件不存在: {json_path}")
        sys.exit(1)
    return json_path


def _keyword_from_json(json_path: str) -> str:
    basename = os.path.basename(json_path)
    name = basename.replace(".json", "")
    parts = name.rsplit("_xianyu_", 1)
    return parts[0] if parts else name


def _screenshots_dir(keyword: str) -> str:
    safe_kw = "".join(c if c.isalnum() or '一' <= c <= '鿿' else '_' for c in keyword)
    return os.path.join(BASE_DIR, "screenshots", safe_kw)


def check_captcha(page) -> bool:
    """多维度判断闲鱼页面异常"""
    url_lower = page.url.lower()
    reason = ""
    try:
        if "goofish.com/item" not in url_lower and "goofish.com/detail" not in url_lower:
            reason = "URL 不是商品详情页"
            raise Exception(reason)

        body_len = page.evaluate("() => (document.body?.innerText || '').length")
        if body_len < 100:
            reason = f"页面文本仅 {body_len} 字符"
            raise Exception(reason)

        title = page.title()
        if any(kw in title for kw in ["验证", "登录", "verify", "login"]):
            reason = f"标题异常: {title}"
            raise Exception(reason)

        return False
    except Exception as e:
        if not reason:
            reason = str(e).replace("Exception: ", "")[:60]

    print()
    print("  ╔══════════════════════════════════════════════╗")
    print(f"  ║  🛑 页面异常 [{reason}]")
    print("  ║     请在浏览器中完成验证/登录                 ║")
    print("  ║     完成后回到此处按 Enter 继续截图            ║")
    print("  ╚══════════════════════════════════════════════╝")
    input("  >>> 按 Enter 继续...")
    print()
    log_ok("已确认，继续截图...")
    return True


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   mtop 签名算法                                   ║
# ╚══════════════════════════════════════════════════════════════════╝

def load_cookies():
    """从 cookies.txt 加载 Cookie"""
    if not os.path.exists(COOKIES_FILE):
        log_err(f"Cookie 文件不存在: {COOKIES_FILE}")
        _log("INFO", "请从浏览器复制完整 Cookie 到 cookies.txt")
        return ""
    with open(COOKIES_FILE, "r", encoding="utf-8") as f:
        return f.read().strip()


def parse_cookie_dict(cookie_str):
    """Cookie 字符串转字典"""
    cookies = {}
    for item in cookie_str.split(";"):
        item = item.strip()
        if "=" in item:
            k, v = item.split("=", 1)
            cookies[k.strip()] = v.strip()
    return cookies


def get_token(cookie_str):
    """从 _m_h5_tk Cookie 提取 token"""
    cookie_dict = parse_cookie_dict(cookie_str)
    h5_tk = cookie_dict.get("_m_h5_tk", "")
    if not h5_tk:
        log_err("Cookie 中缺少 _m_h5_tk，请更新 Cookie")
        return ""
    # 格式: token_timestamp
    token = h5_tk.split("_")[0]
    log_ok(f"Token: {token[:16]}...")
    return token


def compute_sign(token, t, app_key, data):
    """mtop 签名: md5(token + '&' + t + '&' + appKey + '&' + data)"""
    to_sign = f"{token}&{t}&{app_key}&{data}"
    return hashlib.md5(to_sign.encode('utf-8')).hexdigest()


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 1: API 搜索                               ║
# ╚══════════════════════════════════════════════════════════════════╝

def search_once(keyword, page_num, token, cookie_str, session):
    """执行一次搜索请求"""
    t = str(int(time.time() * 1000))

    # 请求数据
    data_obj = {
        "pageNumber": page_num,
        "keyword": keyword,
        "fromFilter": False,
        "rowsPerPage": ROWS_PER_PAGE,
        "sortValue": "",
        "sortField": "",
        "customDistance": "",
        "gps": "",
        "propValueStr": {},
        "customGps": "",
        "searchReqFromPage": "pcSearch",
        "extraFilterValue": "{}",
        "userPositionJson": "{}",
    }
    data_str = json.dumps(data_obj, separators=(',', ':'), ensure_ascii=False)

    # 计算签名
    sign = compute_sign(token, t, APP_KEY, data_str)

    # 构建 URL 参数
    params = {
        "jsv": "2.7.2",
        "appKey": APP_KEY,
        "t": t,
        "sign": sign,
        "v": "1.0",
        "type": "originaljson",
        "accountSite": "xianyu",
        "dataType": "json",
        "timeout": "20000",
        "api": "mtop.taobao.idlemtopsearch.pc.search",
        "sessionOption": "AutoLoginOnly",
        "spm_cnt": "a21ybx.search.0.0",
    }

    url = SEARCH_API + "?" + urllib.parse.urlencode(params)

    headers = {
        "User-Agent": USER_AGENT,
        "Accept": "application/json",
        "Content-Type": "application/x-www-form-urlencoded",
        "Origin": "https://www.goofish.com",
        "Referer": "https://www.goofish.com/",
        "Cookie": cookie_str,
    }

    body = "data=" + urllib.parse.quote(data_str)

    try:
        resp = session.post(url, data=body, headers=headers, timeout=15)

        if resp.status_code != 200:
            log_err(f"HTTP {resp.status_code}")
            return []

        result = resp.json()

        # 检查返回状态
        ret = result.get("ret", [])
        if ret and "FAIL" in str(ret):
            log_err(f"API 错误: {ret}")
            return []

        # 解析商品数据
        data = result.get("data", {})
        items_data = data.get("resultList", [])

        items = []
        for item_wrapper in items_data:
            inner = item_wrapper.get("data", {})
            item_obj = inner.get("item", {})
            main = item_obj.get("main", {})
            ex_content = main.get("exContent", {})
            click_args = main.get("clickParam", {}).get("args", {})

            if not ex_content:
                continue

            # 提取商品信息
            item_id = str(ex_content.get("itemId", click_args.get("id", "")))
            if not item_id:
                continue

            # 标题（优先用 detailParams.title，更完整）
            detail_params = ex_content.get("detailParams", {})
            title = detail_params.get("title", "") or ex_content.get("title", "")
            # 清理标题中的 HTML 标签
            title = title.replace('<em class="highlight">', '').replace('</em>', '')

            # 价格（可能是富文本数组或字符串）
            price_raw = ex_content.get("price", "") or detail_params.get("soldPrice", "") or click_args.get("price", "")
            if isinstance(price_raw, list):
                # 富文本格式: [{"text": "¥"}, {"text": "117"}]
                price_str = "".join(p.get("text", "") for p in price_raw if isinstance(p, dict))
            else:
                price_str = str(price_raw) if price_raw else ""

            # 想要人数（可能是富文本数组或字符串）
            want_raw = ex_content.get("want", "")
            if isinstance(want_raw, list):
                want_text = "".join(w.get("text", "") for w in want_raw if isinstance(w, dict))
            else:
                want_text = str(want_raw) if want_raw else ""
            want_str = f"{want_text}人想要" if want_text and want_text != "0" else ""

            # 卖家信息
            seller_nick = ex_content.get("userNickName", "") or detail_params.get("userNick", "")
            area = ex_content.get("area", "")

            # 图片
            image = ex_content.get("picUrl", "")
            if image and not image.startswith("http"):
                image = "https:" + image if image.startswith("//") else "https://" + image

            # 详情链接
            detail_url = f"https://www.goofish.com/item?id={item_id}"

            items.append({
                "itemId": item_id,
                "title": title,
                "price": price_str,
                "wantCount": want_str,
                "sellerText": f"{seller_nick} · {area}" if seller_nick and area else (seller_nick or area),
                "detailUrl": detail_url,
                "image": image,
                "area": area,
                "sellerNick": seller_nick,
            })

        return items
    except Exception as e:
        log_err(f"请求异常: {e}")
        return []


def step1_api_search(keyword, max_pages=DEFAULT_PAGES):
    """API 搜索多页，返回 JSON 路径"""
    log_step(f"API 搜索: {keyword} | 页数: {max_pages}")

    cookie_str = load_cookies()
    if not cookie_str:
        return None

    token = get_token(cookie_str)
    if not token:
        return None

    import requests
    session = requests.Session()

    all_items = []
    seen_ids = set()

    for page_num in range(1, max_pages + 1):
        log_step(f"第 {page_num} 页")

        items = search_once(keyword, page_num, token, cookie_str, session)
        if not items:
            log_warn(f"第 {page_num} 页无数据，停止")
            break

        new_items = [i for i in items if i["itemId"] not in seen_ids]
        for i in new_items:
            seen_ids.add(i["itemId"])
        all_items.extend(new_items)
        log_ok(f"第 {page_num} 页: {len(new_items)} 条新 (总计 {len(all_items)})")

        if page_num < max_pages:
            delay = random.uniform(1.5, 3.0)
            time.sleep(delay)

    if not all_items:
        log_err("未获取到数据")
        return None

    filename = make_filename(keyword)
    json_path = os.path.join(BASE_DIR, filename)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_items, f, ensure_ascii=False, indent=2)
    log_ok(f"保存 {len(all_items)} 条 → {json_path}")
    return json_path


# ╔══════════════════════════════════════════════════════════════════╗
# ║                  Step 2: 商品详情页截图                           ║
# ╚══════════════════════════════════════════════════════════════════╝

def step2_screenshot(json_path: str, captcha_pause: bool = True, keyword: str = None):
    """读取 JSON，启动浏览器逐条截图"""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        log_err("缺少 playwright，请执行: pip install playwright")
        sys.exit(1)

    exe_path = find_chrome_path()
    if not exe_path or not os.path.exists(exe_path):
        log_err(f"Chrome 不存在: {CHROME_EXE}")
        sys.exit(1)

    log_ok(f"浏览器: {exe_path}")

    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = raw if isinstance(raw, list) else raw.get("data", [])
    log_step(f"共 {len(items)} 个商品待截图")

    if not keyword:
        keyword = _keyword_from_json(json_path)
    screenshots_dir = _screenshots_dir(keyword)
    os.makedirs(screenshots_dir, exist_ok=True)
    done_ids = {f.replace(".png", "") for f in os.listdir(screenshots_dir) if f.endswith(".png")}
    pending = [it for it in items if it.get("itemId", "") not in done_ids]
    skipped = len(items) - len(pending)

    if skipped:
        log_ok(f"跳过 {skipped} 个已有截图，待处理 {len(pending)} 个")
    if not pending:
        log_ok("所有商品已有截图")
        return done_ids

    success = fail = 0
    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                user_data_dir=CHROME_PROFILE, executable_path=exe_path,
                headless=False, viewport={"width": 1280, "height": 900},
                args=["--disable-blink-features=AutomationControlled"])
        except Exception:
            log_err("浏览器启动失败！请关闭占用同 Profile 的 Chrome 窗口")
            sys.exit(1)

        page = context.pages[0] if context.pages else context.new_page()
        log_ok("浏览器已启动，开始截图...")

        for idx, item in enumerate(pending, 1):
            item_id = item.get("itemId", "")
            detail_url = item.get("detailUrl", "")
            if not detail_url:
                log_warn(f"[{idx}/{len(pending)}] 无详情链接，跳过: {item_id}")
                fail += 1
                continue

            snapshot_path = os.path.join(screenshots_dir, f"{item_id}.png")
            try:
                print(f"  [{idx}/{len(pending)}] {item_id} ... ", end="", flush=True)
                page.goto(detail_url, timeout=30000, wait_until="domcontentloaded")
                if captcha_pause and check_captcha(page):
                    page.wait_for_timeout(2000)
                wait_s = random.randint(*PAGE_WAIT)
                page.wait_for_timeout(wait_s * 1000)
                page.screenshot(path=snapshot_path, full_page=False)
                print(f"✓ ({wait_s}s)", flush=True)
                success += 1
                if idx < len(pending):
                    time.sleep(random.uniform(*BETWEEN_WAIT))
            except Exception:
                print("✗", flush=True)
                log_warn(f"[{idx}/{len(pending)}] {item_id} 失败")
                fail += 1

        context.close()

    total = success + skipped + fail
    print()
    log_ok(f"截图完成: 新增 {success}, 跳过 {skipped}, 失败 {fail}, 共 {total}/{len(items)}")
    return done_ids


# ╔══════════════════════════════════════════════════════════════════╗
# ║                   Step 3: JSON → XLSX                            ║
# ╚══════════════════════════════════════════════════════════════════╝

def step3_to_xlsx(json_path: str, existing_ids: set = None, keyword: str = None):
    """将 JSON + 截图合并为 XLSX"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.drawing.image import Image as XlImage
        from openpyxl.utils import get_column_letter
    except ImportError:
        log_err("缺少 openpyxl，请执行: pip install openpyxl Pillow")
        sys.exit(1)

    import urllib.request
    import io as _io

    json_path = resolve_json_path(json_path)
    with open(json_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    items = raw if isinstance(raw, list) else raw.get("data", [])
    if not items:
        log_err("JSON 文件为空")
        return

    log_step(f"数据 {len(items)} 条 → 生成 XLSX ...")

    if not keyword:
        keyword = _keyword_from_json(json_path)
    screenshots_dir = _screenshots_dir(keyword)

    if existing_ids is None:
        existing_ids = set()
        if os.path.exists(screenshots_dir):
            existing_ids = {f.replace(".png", "") for f in os.listdir(screenshots_dir) if f.endswith(".png")}

    wb = Workbook()
    ws = wb.active
    ws.title = "闲鱼搜索结果"

    # 表头样式（闲鱼黄色）
    hdr_fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
    hdr_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    data_font = Font(name="微软雅黑", size=10)
    screenshot_col = len(HEADERS)
    cover_col = 10  # 商品图片列
    img_ok = img_miss = cover_ok = cover_miss = 0

    for row_idx, item in enumerate(items, 2):
        item_id = item.get("itemId", "")
        status = "已爬取" if item_id in existing_ids else "新增"
        row_data = [
            row_idx - 1,
            item.get("title", ""),
            item_id,
            item.get("price", ""),
            item.get("wantCount", ""),
            item.get("sellerNick", ""),
            item.get("area", ""),
            status,
            item.get("detailUrl", ""),
            "",  # 商品图片
            "",  # 截图
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            align_h = "center" if col in (1, 3, 4, 5, 8) else "left"
            cell.alignment = Alignment(horizontal=align_h, vertical="center", wrap_text=(col in (2, 9)))

        ws.row_dimensions[row_idx].height = max(IMG_HEIGHT_PT, COVER_HEIGHT_PT) + 8

        # 商品图片
        cover_url = item.get("image", "")
        if cover_url:
            try:
                if cover_url.startswith("//"):
                    cover_url = "https:" + cover_url
                req = urllib.request.Request(cover_url, headers={
                    "User-Agent": USER_AGENT, "Referer": "https://www.goofish.com/"})
                with urllib.request.urlopen(req, timeout=15) as resp:
                    img_data = resp.read()
                cover_img = XlImage(_io.BytesIO(img_data))
                cover_img.width = COVER_WIDTH_PT
                cover_img.height = COVER_HEIGHT_PT
                ws.add_image(cover_img, f"{get_column_letter(cover_col)}{row_idx}")
                cover_ok += 1
            except Exception:
                ws.cell(row=row_idx, column=cover_col, value=cover_url)
                cover_miss += 1
        else:
            cover_miss += 1

        # 截图
        img_path = os.path.join(screenshots_dir, f"{item_id}.png")
        if os.path.exists(img_path):
            try:
                img = XlImage(img_path)
                img.width = IMG_WIDTH_PT
                img.height = IMG_HEIGHT_PT
                ws.add_image(img, f"{get_column_letter(screenshot_col)}{row_idx}")
                img_ok += 1
            except Exception:
                ws.cell(row=row_idx, column=screenshot_col, value="(截图加载失败)")
                img_miss += 1
        else:
            ws.cell(row=row_idx, column=screenshot_col, value="(无截图)")
            img_miss += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(items) + 1}"

    out_path = json_path.rsplit(".", 1)[0] + ".xlsx"
    wb.save(out_path)

    print()
    log_ok(f"XLSX 已保存: {out_path}")
    log_ok(f"数据 {len(items)} 条 | 商品图 {cover_ok}/{len(items)} | 截图 {img_ok}/{len(items)}")
    if cover_miss:
        log_warn(f"{cover_miss} 条商品图下载失败")
    if img_miss:
        log_warn(f"{img_miss} 条缺少截图")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                         主入口                                   ║
# ╚══════════════════════════════════════════════════════════════════╝

def main():
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    print("=" * 56)
    print("    闲鱼商品搜索采集工具 - API 协议版")
    print("=" * 56)

    args = sys.argv[1:]

    pages = DEFAULT_PAGES
    if "--pages" in args:
        pi = args.index("--pages")
        if pi + 1 < len(args):
            pages = int(args[pi + 1])
    captcha_pause = "--no-captcha" not in args

    if "--auto" in args:
        # 模式3: API搜索 → 截图 → XLSX
        idx = args.index("--auto")
        keyword = args[idx + 1] if idx + 1 < len(args) else ""
        if not keyword or keyword.startswith("--"):
            keyword = input("请输入搜索关键词: ").strip()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)

        json_path = step1_api_search(keyword, max_pages=pages)
        if not json_path:
            sys.exit(1)
        print()
        log_step("继续截图...")
        existing_ids = step2_screenshot(json_path, captcha_pause=captcha_pause, keyword=keyword)
        print()
        log_step("继续生成 XLSX...")
        step3_to_xlsx(json_path, existing_ids=existing_ids, keyword=keyword)

    elif "--from-json" in args:
        # 模式2: JSON → 截图 → XLSX
        idx = args.index("--from-json")
        json_file = args[idx + 1] if idx + 1 < len(args) else ""
        if not json_file:
            log_err("请指定 JSON 文件路径")
            sys.exit(1)
        keyword = _keyword_from_json(json_file)
        existing_ids = step2_screenshot(json_file, captcha_pause=captcha_pause, keyword=keyword)
        print()
        step3_to_xlsx(json_file, existing_ids=existing_ids, keyword=keyword)

    else:
        # 模式1: API 搜索 → JSON
        keyword = args[0] if args else ""
        if not keyword or keyword.startswith("--"):
            keyword = input("请输入搜索关键词: ").strip()
        if not keyword:
            log_err("关键词不能为空")
            sys.exit(1)
        step1_api_search(keyword, max_pages=pages)


if __name__ == "__main__":
    main()
